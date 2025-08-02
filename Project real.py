import pandas as pd
import logging
import openpyxl
import os
from openpyxl.styles import Alignment, Font 
from openpyxl.utils import get_column_letter


siswa = []  # List untuk menyimpan data siswa
mata_pelajaran = ["agama", "ppkn", "bind", "mtk", "pjok", "seni_rupa", "plbj"]  # List mata pelajaran

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Fungsi untuk menambahkan siswa beserta nilai dan ketidakhadiran
def tambah_siswa():
    """Fungsi untuk menambahkan siswa beserta nilai dan ketidakhadiran"""
    logging.info("Fungsi tambah siswa dipanggil.")
    print("\n=== Tambah Siswa ===")
    
    try:
        nis = input("Masukkan NIS: ")
        nama = input("Masukkan Nama Siswa: ")
        kelas = input("Masukkan Kelas: ")
        
        # Input nilai dengan validasi
        print("\n--- Input Nilai ---")
        agama = int(input("Masukkan nilai Agama (0-100): "))
        ppkn = int(input("Masukkan nilai PPKn (0-100): "))
        bind = int(input("Masukkan nilai Bahasa Indonesia (0-100): "))
        mtk = int(input("Masukkan nilai Matematika (0-100): "))
        pjok = int(input("Masukkan nilai PJOK (0-100): "))
        seni_rupa = int(input("Masukkan nilai Seni Rupa (0-100): "))
        plbj = int(input("Masukkan nilai PLBJ (0-100): "))
        
        # Input ketidakhadiran
        print("\n--- Input Ketidakhadiran ---")
        sakit = int(input("Masukkan jumlah hari sakit: "))
        izin = int(input("Masukkan jumlah hari izin: "))
        alpa = int(input("Masukkan jumlah hari alpa: "))
        
        # Input ekstrakurikuler
        print("\n--- Input Ekstrakurikuler ---")
        print("Pilihan: Seni Lukis, Pramuka, Seni Tari, Pencak Silat")
        ekstra_input = input("Masukkan ekstrakurikuler (pisahkan dengan koma): ")
        ekstra = [e.strip() for e in ekstra_input.split(",") if e.strip()]

        # Menambahkan data siswa ke dalam daftar
        siswa.append({
            "nis": nis,
            "nama": nama,
            "kelas": kelas,
            "nilai": {
                "agama": agama,
                "ppkn": ppkn,
                "bind": bind,
                "mtk": mtk,
                "pjok": pjok,
                "seni_rupa": seni_rupa,
                "plbj": plbj
            },
            "ketidakhadiran": {
                "sakit": sakit,
                "izin": izin,
                "alpa": alpa
            },
            "ekstra_kurikuler": ekstra
        })
        print(f"\nSiswa {nama} berhasil ditambahkan!")
        
    except ValueError:
        print("Error: Harap masukkan angka yang valid untuk nilai dan ketidakhadiran.")
    except Exception as e:
        print(f"Error: {e}")

#tampilkan data siswa
def tampilkan_data():
    """Fungsi untuk menampilkan data siswa"""
    logging.info("Fungsi Menampilkan Data Siswa.")
    if not siswa:
        print("\nBelum ada data siswa.")
        return
    
    print("\n" + "="*60)
    print("DATA SISWA SD NEGERI 58")
    print("="*60)
    
    for i, s in enumerate(siswa, start=1):
        print(f"\n{i}. NIS: {s['nis']} | Nama: {s['nama']} | Kelas: {s['kelas']}")
        print("   " + "-"*50)
        
        # Tampilkan nilai
        print("   📚 NILAI:")
        for mapel, nilai in s["nilai"].items():
            mapel_display = mapel.replace("_", " ").title()
            print(f"     • {mapel_display:<20}: {nilai}")
        
        # Hitung rata-rata
        rata_rata = sum(s["nilai"].values()) / len(s["nilai"])
        print(f"     • {'Rata-rata':<20}: {rata_rata:.2f}")
        
        # Tampilkan ketidakhadiran
        print("   📅 KETIDAKHADIRAN:")
        for jenis, jumlah in s["ketidakhadiran"].items():
            print(f"     • {jenis.capitalize():<20}: {jumlah} hari")
        
        # Tampilkan ekstrakurikuler
        print("   🏆 EKSTRAKURIKULER:")
        if s["ekstra_kurikuler"]:
            for ekstra in s["ekstra_kurikuler"]:
                print(f"•                 python --version                python --version{ekstra}")
        else:
            print("• Tidak ada")
        
        print("   " + "-"*50)

#urutkan siswa berdasarkan NIS
def urutkan_siswa():
    """Fungsi untuk mengurutkan siswa berdasarkan NIS"""
    logging.info("Fungsi Urutkan Siswa Dipanggil.")
    if not siswa:
        print("\nBelum ada data siswa untuk diurutkan.")
        return
    
    siswa.sort(key=lambda x: x["nis"])
    print("\n✅ Data siswa berhasil diurutkan berdasarkan NIS!")

# Fungsi untuk menambahkan atau mengubah nilai siswa
def tambah_nilai():
    """Fungsi untuk menambahkan atau mengubah nilai siswa"""
    logging.info("Fungsi tambah nilai dipanggil.")
    
    if not siswa:
        print("\nBelum ada data siswa.")
        return
    nis = input("Masukkan NIS siswa: ")
    siswa_ditemukan = next((s for s in siswa if s["nis"] == nis), None)
    if not siswa_ditemukan:
        print("❌ Siswa dengan NIS tersebut tidak ditemukan.")
        return
    
# Tambahan: Edit nama siswa
    print(f"Nama siswa saat ini: {siswa_ditemukan['nama']}")
    ubah_nama = input("Apakah ingin mengubah nama siswa? (y/n): ").lower()
    if ubah_nama == "y":
        nama_baru = input("Masukkan nama baru: ")
        siswa_ditemukan["nama"] = nama_baru
        print("Nama siswa berhasil diubah.")    
    
    print(f"\n📝 Mengubah nilai untuk: {siswa_ditemukan['nama']}")
    print("Mata pelajaran yang tersedia:")
    
    for i, mapel in enumerate(mata_pelajaran):
        mapel_display = mapel.replace("_", " ").title()
        current_value = siswa_ditemukan["nilai"].get(mapel, "Belum ada")
        print(f"{i + 1}. {mapel_display} (Nilai saat ini: {current_value})")
        
    
    try:
        idx = int(input("Pilih nomor mata pelajaran: ")) - 1
        if 0 <= idx < len(mata_pelajaran):
            mapel = mata_pelajaran[idx]
            nilai_lama = siswa_ditemukan["nilai"].get(mapel, None)
            print(f"Nilai {mapel} saat ini: {nilai_lama}")
            konfirmasi = input(f"Apakah ingin mengubah nilai {mapel}? (y/n): ").lower()
            if konfirmasi == "y":
                nilai = float(input(f"Masukkan nilai baru untuk {mapel}: "))
                siswa_ditemukan["nilai"][mapel] = nilai
                print("Nilai berhasil diubah.")
            else:
                print("Nilai tidak diubah.")
    except ValueError:
        print("❌ Input tidak valid. Harap masukkan angka.")

#Load data dari Excel
def load_data_from_excel():
    """Fungsi untuk memuat data dari Excel  Per Kelas"""
    global siswa
    logging.info("Fungsi Memuat Data dari Excel.")
    
    file_list = [
        'data_Kelas1A_updated.xlsx', 'data_Kelas1B_updated.xlsx',
        'data_Kelas2A_updated.xlsx', 'data_Kelas2B_updated.xlsx',
        'data_Kelas3A_updated.xlsx', 'data_Kelas3B_updated.xlsx',
        'data_Kelas4A_updated.xlsx', 'data_Kelas4B_updated.xlsx',
        'data_Kelas5A_updated.xlsx', 'data_Kelas5B_updated.xlsx',
        'data_Kelas6A_updated.xlsx', 'data_Kelas6B_updated.xlsx',
    ]
    siswa.clear()  # Bersihkan list siswa sebelum memuat baru
    total_berhasil = 0

    for file in file_list:
        if not os.path.isfile(file):
            print(f"❌ File {file} tidak ditemukan, dilewati.")
            continue

        try:
            df = pd.read_excel(file)

            kolom_minimal = ["NAMA SISWA", "NIS", "KELAS"]
            for kolom in kolom_minimal:
                if kolom not in df.columns:
                    raise KeyError(f"Kolom '{kolom}' wajib ada.")

            for _, row in df.iterrows():
                try:
                    # Ambil nilai aman
                    nama = str(row.get("NAMA SISWA", "")).strip()
                    nis = str(row.get("NIS", "")).strip()
                    kelas = str(row.get("KELAS", "")).strip()

                    nilai = {
                        "agama": int(row.get("AGAMA", 0)),
                        "ppkn": int(row.get("PPKN", 0)),
                        "bind": int(row.get("BIND", 0)),
                        "mtk": int(row.get("MTK", 0)),
                        "pjok": int(row.get("PJOK", 0)),
                        "seni_rupa": int(row.get("SENI RUPA", 0)),
                        "plbj": int(row.get("PLBJ", 0))
                    }

                    ketidakhadiran = {
                        "sakit": int(row.get("SAKIT", 0)),
                        "izin": int(row.get("IZIN", 0)),
                        "alpa": int(row.get("ALPA", row.get("ALPHA", 0)))  # Tangani jika kolom ALPA tertulis ALPHA
                    }

                    ekstra_raw = str(row.get("EKSTRA KURIKULER", "")).strip()
                    ekstra = [e.strip() for e in ekstra_raw.split(",") if e.strip()] if ekstra_raw else []

                    siswa.append({
                        "nama": nama,
                        "nis": nis,
                        "kelas": kelas,
                        "nilai": nilai,
                        "ketidakhadiran": ketidakhadiran,
                        "ekstra_kurikuler": ekstra
                    })

                    total_berhasil += 1

                except Exception as row_error:
                    print(f"⚠️  Gagal membaca 1 baris di file {file}: {row_error}")
                    continue

            print(f"✅ Data siswa dari {file} berhasil dimuat.")

        except KeyError as ke:
            print(f"❌ Struktur kolom tidak sesuai di file {file}: {ke}")
        except Exception as e:
            print(f"❌ Terjadi kesalahan saat membaca {file}: {e}")

    print(f"📦 Total siswa yang dimuat: {total_berhasil}")


# Fungsi untuk mengekspor data siswa ke file Excel per kelas
def export_per_kelas():
    if not siswa:
        print("Belum ada data siswa.")
        return

    # Kelompokkan siswa berdasarkan kelas
    kelas_dict = {}
    for s in siswa:
        kelas = s["kelas"]
        if kelas not in kelas_dict:
            kelas_dict[kelas] = []
        kelas_dict[kelas].append(s)
        
    for kelas, data_siswa in kelas_dict.items():    
        data = []
        for idx, s in enumerate(data_siswa, start=1):
            row ={
        "NO": idx,
        "NAMA SISWA": s["nama"],
        "NIS": s["nis"],
        "KELAS": s["kelas"],
        "AGAMA": s["nilai"]["agama"],
        "PPKN": s["nilai"]["ppkn"],
        "BIND": s["nilai"]["bind"],
        "MTK": s["nilai"]["mtk"],
        "PJOK": s["nilai"]["pjok"],
        "SENI RUPA": s["nilai"]["seni_rupa"],
        "PLBJ": s["nilai"]["plbj"],
        "RATA-RATA": round(sum(s["nilai"].values()) / len(s["nilai"]), 2),
        "SAKIT": s["ketidakhadiran"]["sakit"],
        "IZIN": s["ketidakhadiran"]["izin"],
        "ALPA": s["ketidakhadiran"]["alpa"],
        "EKSTRA KURIKULER": ", ".join(s["ekstra_kurikuler"]) if s["ekstra_kurikuler"] else "Tidak ada"
    }
    data.append(row)
        
# Buat DataFrame dan simpan ke Excel
    df_new = pd.DataFrame(data)
    file_name = f"data_kelas_{kelas}.xlsx"
    df_new.to_excel(file_name, index=False)
    print(f"Data kelas {kelas} berhasil disimpan ke {file_name}")

    try:         
        # Styling dengan openpyxl
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # style data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # kolom lebar di excel
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 5, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_name)
        print(f"✅ Data kelas {kelas} berhasil disimpan ke {file_name} dengan styling!")
    except Exception as e:
        print(f"❌ Error saat menyimpan data: {e}")
        
# Fungsi untuk mencari siswa berdasarkan NIS atau nama
def cari_siswa():
    """Fungsi untuk mencari siswa berdasarkan NIS atau nama"""
    if not siswa:
        print("\nBelum ada data siswa.")
        return
    
    kata_kunci = input("Masukkan NIS atau nama siswa: ").lower()
    hasil_pencarian = []
    
    for s in siswa:
        if kata_kunci in s["nis"].lower() or kata_kunci in s["nama"].lower():
            hasil_pencarian.append(s)
    
    if hasil_pencarian:
        print(f"\n🔍 Ditemukan {len(hasil_pencarian)} siswa:")
        for i, s in enumerate(hasil_pencarian, 1):
            print(f"{i}. NIS: {s['nis']} | Nama: {s['nama']} | Kelas: {s['kelas']}")
    else:
        print("❌ Siswa tidak ditemukan.")

def statistik_kelas():
    """Fungsi untuk menampilkan statistik kelas"""
    if not siswa:
        print("\nBelum ada data siswa.")
        return
    
    print("\n📊 STATISTIK KELAS")
    print("="*30)
    
    # Statistik per mata pelajaran
    for mapel in mata_pelajaran:
        nilai_mapel = [s["nilai"][mapel] for s in siswa]
        if nilai_mapel:
            rata_rata = sum(nilai_mapel) / len(nilai_mapel)
            nilai_tertinggi = max(nilai_mapel)
            nilai_terendah = min(nilai_mapel)
            
            mapel_display = mapel.replace("_", " ").title()
            print(f"\n{mapel_display}:")
            print(f"  • Rata-rata: {rata_rata:.2f}")
            print(f"  • Tertinggi: {nilai_tertinggi}")
            print(f"  • Terendah: {nilai_terendah}")
    
    # Statistik kehadiran
    total_sakit = sum(s["ketidakhadiran"]["sakit"] for s in siswa)
    total_izin = sum(s["ketidakhadiran"]["izin"] for s in siswa)
    total_alpa = sum(s["ketidakhadiran"]["alpa"] for s in siswa)
    
    print(f"\nStatistik Kehadiran:")
    print(f"  • Total hari sakit: {total_sakit}")
    print(f"  • Total hari izin: {total_izin}")
    print(f"  • Total hari alpa: {total_alpa}")

# Fungsi untuk mencetak rapor per siswa ke file Excel
def cetak_rapor_per_siswa():
    if not siswa:
        print("Belum ada data siswa.")
        return

    for s in siswa:
        file_name = f"RAPOR_{s['nama'].replace(' ', '_')}_{s['nis']}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapor"

        # Judul
        ws.merge_cells('A1:E1')
        ws['A1'] = "RAPOR SISWA SD NEGERI 58"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Identitas
        ws['A3'] = "Nama"
        ws['B3'] = s['nama']
        ws['A4'] = "NIS"
        ws['B4'] = s['nis']
        ws['A5'] = "Kelas"
        ws['B5'] = s['kelas']

        # Nilai
        ws['A7'] = "Mata Pelajaran"
        ws['B7'] = "Nilai"
        ws['A7'].font = ws['B7'].font = Font(bold=True)
        ws['A7'].alignment = ws['B7'].alignment = Alignment(horizontal="center")

        row_num = 8
        for mapel, nilai in s["nilai"].items():
            ws[f'A{row_num}'] = mapel.replace("_", " ").title()
            ws[f'B{row_num}'] = nilai
            row_num += 1

        # Rata-rata
        ws[f'A{row_num}'] = "Rata-rata"
        ws[f'B{row_num}'] = round(sum(s["nilai"].values()) / len(s["nilai"]), 2)
        ws[f'A{row_num}'].font = Font(bold=True)
        ws[f'B{row_num}'].font = Font(bold=True)
        row_num += 2

        # Kehadiran
        ws[f'A{row_num}'] = "Kehadiran"
        ws[f'A{row_num}'].font = Font(bold=True)
        row_num += 1
        for jenis, jumlah in s["ketidakhadiran"].items():
            ws[f'A{row_num}'] = jenis.capitalize()
            ws[f'B{row_num}'] = jumlah
            row_num += 1

        # Ekstrakurikuler
        row_num += 1
        ws[f'A{row_num}'] = "Ekstrakurikuler"
        ws[f'A{row_num}'].font = Font(bold=True)
        ws[f'B{row_num}'] = ", ".join(s["ekstra_kurikuler"]) if s["ekstra_kurikuler"] else "Tidak ada"

        # Atur lebar kolom
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_name)
        print(f"✅ Rapor untuk {s['nama']} ({s['nis']}) berhasil disimpan ke {file_name}")

#fungsi menu utama
def menu():
    """Fungsi menu utama"""
    
    # Tampilan menu utama
    logging.info("Menampilkan menu utama.")
    while True:
        print("\n" + "="*50)
        print("\033[4m\033[1m\033[34m\033[44m\033[1;94m\n🏫 SISTEM MANAJEMEN SISWA SD NEGERI 58 🏫\033[0m \033[1m".center(20))
        print("="*50)
        print("1. 👤 Tambah Siswa")
        print("2. 📋 Tampilkan Data Siswa")
        print("3. 🔢 Urutkan Siswa Berdasarkan NIS")
        print("4. ✏️  Tambah/Ubah Nilai Dan Nama Siswa")
        print("5. 🔍 Cari Siswa")
        print("6. 📊 Statistik Kelas")
        print("7. 💾 Simpan Data ke Excel")
        print("8. 📁 Muat Data dari Excel")
        print("9. 🚪 Keluar")
        print("10. 🖨️  Cetak Rapor Per Siswa")
        print("="*50)
        
        pilihan = input("Pilih menu (1-9): ").strip()
        
        if pilihan == "1":
            tambah_siswa()
        elif pilihan == "2":
            tampilkan_data()
        elif pilihan == "3":
            urutkan_siswa()
        elif pilihan == "4":
            tambah_nilai()
        elif pilihan == "5":
            cari_siswa()
        elif pilihan == "6":
            statistik_kelas()
        elif pilihan == "7":
            export_per_kelas()
        elif pilihan == "8":
            load_data_from_excel()
        elif pilihan == "9":
            print("\n Terima kasih telah menggunakan Sistem Manajemen Siswa SD Negeri 58! 🙏")
            break
        elif pilihan == "10":
            cetak_rapor_per_siswa()
        else:
            print("❌ Pilihan tidak valid. Silakan pilih menu 1-9.")
        
        # Pause untuk membaca output
        input("\nTekan Enter untuk melanjutkan...")

if __name__ == "__main__":
    menu()